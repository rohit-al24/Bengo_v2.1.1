import io
import random

import openpyxl
from django.http import HttpResponse
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated, IsAdminUser
from rest_framework.response import Response
from rest_framework.views import APIView

from . import models, serializers


# ── Helpers ────────────────────────────────────────────────────────────────────

def _is_admin(user):
    roles = list(user.roles.values_list('name', flat=True)) if hasattr(user, 'roles') else []
    return user.is_staff or 'admin' in roles


# ── Admin: Stories ──────────────────────────────────────────────────────────────

class AdminStoryView(APIView):
    """GET = list all stories; POST = create story."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        qs = models.RolePlayStory.objects.all()
        return Response(serializers.RolePlayStoryListSerializer(qs, many=True).data)

    def post(self, request):
        if not _is_admin(request.user):
            return Response({'detail': 'Forbidden'}, status=403)
        s = serializers.RolePlayStoryListSerializer(data=request.data)
        s.is_valid(raise_exception=True)
        s.save()
        return Response(s.data, status=201)


class AdminStoryDetailView(APIView):
    permission_classes = [IsAuthenticated]

    def get_object(self, pk):
        return models.RolePlayStory.objects.get(pk=pk)

    def get(self, request, pk):
        obj = self.get_object(pk)
        return Response(serializers.RolePlayStorySerializer(obj).data)

    def patch(self, request, pk):
        if not _is_admin(request.user):
            return Response({'detail': 'Forbidden'}, status=403)
        obj = self.get_object(pk)
        s = serializers.RolePlayStoryListSerializer(obj, data=request.data, partial=True)
        s.is_valid(raise_exception=True)
        s.save()
        return Response(s.data)

    def delete(self, request, pk):
        if not _is_admin(request.user):
            return Response({'detail': 'Forbidden'}, status=403)
        obj = self.get_object(pk)
        obj.delete()
        return Response(status=204)


class AdminStoryTemplateView(APIView):
    """Download the Excel import template."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'RolePlay Stories'

        headers = [
            'Story_Title', 'Category', 'Difficulty', 'Cover_Emoji',
            'Char_Name', 'Char_Emoji', 'Char_Order',
            'Dialogue_Order', 'Japanese', 'Romaji', 'English', 'Emotion', 'Pause_MS',
        ]
        ws.append(headers)

        # Sample row 1
        ws.append([
            'Restaurant Conversation', 'Daily Life', 'easy', '🍜',
            'Customer', '👤', 1,
            1, 'すみません、席はありますか？', 'Sumimasen, seki wa arimasu ka?',
            'Excuse me, do you have a seat?', 'polite', 1000,
        ])
        # Sample row 2
        ws.append([
            'Restaurant Conversation', 'Daily Life', 'easy', '🍜',
            'Waiter', '🧑‍🍳', 2,
            2, 'はい、こちらへどうぞ。', 'Hai, kochira e dōzo.',
            'Yes, right this way.', 'polite', 1000,
        ])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        resp['Content-Disposition'] = 'attachment; filename="roleplay_template.xlsx"'
        return resp


class AdminStoryImportView(APIView):
    """POST with multipart file to import an Excel sheet."""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if not _is_admin(request.user):
            return Response({'detail': 'Forbidden'}, status=403)

        file = request.FILES.get('file')
        exam_id = request.data.get('exam_id')
        if not file:
            return Response({'detail': 'No file provided.'}, status=400)
        if not exam_id:
            return Response({'detail': 'No exam selected.'}, status=400)

        # Retrieve selected Exam
        from apps.courses.models import Exam
        try:
            exam = Exam.objects.get(pk=exam_id)
        except Exam.DoesNotExist:
            return Response({'detail': 'Selected exam does not exist.'}, status=400)

        try:
            wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
        except Exception:
            return Response({'detail': 'Could not read Excel file.'}, status=400)

        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))

        REQUIRED = ['Story_Title', 'Char_Name', 'Char_Order', 'Dialogue_Order', 'Japanese']
        header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        col_map = {str(h).strip(): i for i, h in enumerate(header_row) if h}

        missing = [c for c in REQUIRED if c not in col_map]
        if missing:
            return Response({'detail': f"Missing columns: {missing}"}, status=400)

        def cell(row, name, default=''):
            idx = col_map.get(name)
            val = row[idx] if idx is not None and idx < len(row) else None
            return str(val).strip() if val is not None else default

        stories_created = 0
        dialogues_created = 0
        errors = []

        # Group rows by story title
        story_map = {}
        for row in rows:
            t = cell(row, 'Story_Title')
            if not t:
                continue
            story_map.setdefault(t, []).append(row)

        for title, story_rows in story_map.items():
            first = story_rows[0]
            story, created = models.RolePlayStory.objects.get_or_create(
                title=title,
                defaults={
                    'category':    cell(first, 'Category', 'General'),
                    'exam':        exam,
                    'difficulty':  cell(first, 'Difficulty', 'easy').lower(),
                    'cover_emoji': cell(first, 'Cover_Emoji', '📖'),
                    'is_published': True,
                },
            )
            if created:
                stories_created += 1

            # Gather characters
            char_map = {}
            for row in story_rows:
                cn = cell(row, 'Char_Name')
                if not cn or cn in char_map:
                    continue
                try:
                    order = int(float(cell(row, 'Char_Order', '1')))
                except ValueError:
                    order = 1
                char, _ = models.RolePlayCharacter.objects.get_or_create(
                    story=story,
                    name=cn,
                    defaults={
                        'emoji':         cell(row, 'Char_Emoji', '👤'),
                        'display_order': order,
                    },
                )
                char_map[cn] = char

            # Build dialogues
            for row in story_rows:
                cn = cell(row, 'Char_Name')
                jp = cell(row, 'Japanese')
                if not jp or cn not in char_map:
                    continue
                try:
                    d_order = int(float(cell(row, 'Dialogue_Order', '1')))
                    pause   = int(float(cell(row, 'Pause_MS', '1000')))
                except ValueError:
                    d_order, pause = 1, 1000

                emotion_raw = cell(row, 'Emotion', 'neutral').lower()
                valid_emotions = ['happy', 'sad', 'angry', 'polite', 'neutral', 'serious', 'excited']
                emotion = emotion_raw if emotion_raw in valid_emotions else 'neutral'

                models.RolePlayDialogue.objects.get_or_create(
                    story=story,
                    display_order=d_order,
                    defaults={
                        'character': char_map[cn],
                        'japanese':  jp,
                        'romaji':    cell(row, 'Romaji', ''),
                        'english':   cell(row, 'English', ''),
                        'emotion':   emotion,
                        'pause_ms':  pause,
                    },
                )
                dialogues_created += 1

        return Response({
            'stories_created':   stories_created,
            'dialogues_created': dialogues_created,
            'errors':            errors,
        }, status=201)


# ── Public: Story list ──────────────────────────────────────────────────────────

class PublicStoryListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        qs = models.RolePlayStory.objects.filter(is_published=True)
        return Response(serializers.RolePlayStoryListSerializer(qs, many=True).data)


class PublicStoryDetailView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, pk):
        try:
            obj = models.RolePlayStory.objects.get(pk=pk, is_published=True)
        except models.RolePlayStory.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=404)
        return Response(serializers.RolePlayStorySerializer(obj).data)


# ── Rooms ──────────────────────────────────────────────────────────────────────

class RoomListCreateView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        visibility = request.query_params.get('visibility')
        qs = models.RolePlayRoom.objects.filter(status='waiting')
        if visibility:
            qs = qs.filter(visibility=visibility)
        return Response(serializers.RolePlayRoomSerializer(qs, many=True).data)

    def post(self, request):
        data = request.data.copy()
        # Creator joins automatically
        room = models.RolePlayRoom.objects.create(
            creator=request.user,
            visibility=data.get('visibility', 'public'),
            max_players=int(data.get('max_players', 4)),
        )
        models.RolePlayMember.objects.create(room=room, user=request.user, is_creator=True)
        return Response(serializers.RolePlayRoomSerializer(room).data, status=201)


class RoomDetailView(APIView):
    permission_classes = [IsAuthenticated]

    def _get_room(self, code):
        try:
            return models.RolePlayRoom.objects.get(room_code=code.upper())
        except models.RolePlayRoom.DoesNotExist:
            return None

    def get(self, request, code):
        room = self._get_room(code)
        if not room:
            return Response({'detail': 'Room not found.'}, status=404)
        return Response(serializers.RolePlayRoomSerializer(room).data)


class RoomJoinView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, code):
        try:
            room = models.RolePlayRoom.objects.get(room_code=code.upper())
        except models.RolePlayRoom.DoesNotExist:
            return Response({'detail': 'Room not found.'}, status=404)

        if room.status != 'waiting':
            return Response({'detail': 'Room already started.'}, status=400)

        if room.members.count() >= room.max_players:
            return Response({'detail': 'Room is full.'}, status=400)

        member, created = models.RolePlayMember.objects.get_or_create(
            room=room, user=request.user,
        )
        return Response(serializers.RolePlayRoomSerializer(room).data,
                        status=201 if created else 200)


class RoomSpinView(APIView):
    """Creator spins — backend randomly picks a published story."""
    permission_classes = [IsAuthenticated]

    def post(self, request, code):
        try:
            room = models.RolePlayRoom.objects.get(room_code=code.upper())
        except models.RolePlayRoom.DoesNotExist:
            return Response({'detail': 'Room not found.'}, status=404)

        if room.creator != request.user:
            return Response({'detail': 'Only the creator can spin.'}, status=403)
        if room.status != 'waiting':
            return Response({'detail': 'Already spun.'}, status=400)

        stories = list(models.RolePlayStory.objects.filter(is_published=True))
        if not stories:
            return Response({'detail': 'No published stories available.'}, status=400)

        story = random.choice(stories)
        room.story = story
        room.status = 'active'
        room.save()

        return Response({
            'story_id':    story.id,
            'story_title': story.title,
            'story_emoji': story.cover_emoji,
            'story_jlpt':  story.jlpt_level,
            'room':        serializers.RolePlayRoomSerializer(room).data,
        })


class RoomSelectCharacterView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, code):
        try:
            room = models.RolePlayRoom.objects.get(room_code=code.upper())
        except models.RolePlayRoom.DoesNotExist:
            return Response({'detail': 'Room not found.'}, status=404)

        char_id = request.data.get('character_id')
        if not char_id:
            return Response({'detail': 'character_id required.'}, status=400)

        try:
            character = models.RolePlayCharacter.objects.get(pk=char_id, story=room.story)
        except models.RolePlayCharacter.DoesNotExist:
            return Response({'detail': 'Character not found in this story.'}, status=404)

        # Prevent two members picking same character
        if room.members.filter(character=character).exclude(user=request.user).exists():
            return Response({'detail': 'Character already taken.'}, status=409)

        member = models.RolePlayMember.objects.get(room=room, user=request.user)
        member.character = character
        member.save()

        return Response(serializers.RolePlayMemberSerializer(member).data)


class RoomSubmitLineView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, code):
        try:
            room = models.RolePlayRoom.objects.get(room_code=code.upper())
        except models.RolePlayRoom.DoesNotExist:
            return Response({'detail': 'Room not found.'}, status=404)

        dialogue_id = request.data.get('dialogue_id')
        raw_correct = request.data.get('correct', False)
        raw_passed  = request.data.get('passed', False)
        if isinstance(raw_correct, str):
            correct = raw_correct.lower() == 'true'
        else:
            correct = bool(raw_correct)
        if isinstance(raw_passed, str):
            passed = raw_passed.lower() == 'true'
        else:
            passed = bool(raw_passed)
        score       = float(request.data.get('score', 0.0))

        try:
            dialogue = models.RolePlayDialogue.objects.get(pk=dialogue_id, story=room.story)
        except models.RolePlayDialogue.DoesNotExist:
            return Response({'detail': 'Dialogue not found.'}, status=404)

        member = models.RolePlayMember.objects.get(room=room, user=request.user)

        result, created = models.RolePlayLineResult.objects.get_or_create(
            room=room, member=member, dialogue=dialogue,
            defaults={'correct': correct, 'score': score},
        )

        recording_file = request.FILES.get('recording')
        if recording_file and (correct or passed):
            if result.recording:
                result.recording.delete(save=False)
            result.recording.save(recording_file.name, recording_file, save=False)

        if correct:
            result.correct = True
            result.score = float(score)
            member.score = float(member.score) + score
            member.save()

        if recording_file and (correct or passed):
            result.save()
        elif created:
            result.save()

        if correct or passed:
            # Advance room turn
            room.current_dialogue_index += 1
            # Check if finished
            dialogue_count = room.story.dialogues.count()
            if room.current_dialogue_index >= dialogue_count:
                room.status = 'finished'
            room.save()

        return Response({
            'correct':     result.correct,
            'score':       result.score,
            'total_score': member.score,
            'current_dialogue_index': room.current_dialogue_index,
            'room_status': room.status,
            'recording_url': result.recording.url if result.recording else None,
        })


# ── History ────────────────────────────────────────────────────────────────────

class MyHistoryView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        memberships = models.RolePlayMember.objects.filter(
            user=request.user,
            room__status='finished',
        ).select_related('room', 'room__story').prefetch_related('line_results')
        return Response(serializers.RolePlayHistorySerializer(memberships, many=True).data)
