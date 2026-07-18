import io
from typing import List, Dict, Any

from django.db import transaction
from openpyxl import Workbook, load_workbook

from apps.accounts.models import Role, User, UserRole
from .models import Institution, MentorAssignment


def build_template_bytes(model_name: str, headers: List[str]) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = model_name
    worksheet.append(headers)

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def _read_rows(file_obj) -> List[Dict[str, Any]]:
    workbook = load_workbook(file_obj, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(cell).strip() if cell is not None else '' for cell in rows[0]]
    parsed_rows = []
    for row in rows[1:]:
        if not any(cell not in (None, '') for cell in row):
            continue
        normalized_row = {}
        for index, header in enumerate(headers):
            value = row[index] if index < len(row) else ''
            normalized_row[header] = '' if value is None else str(value).strip()
        parsed_rows.append(normalized_row)
    return parsed_rows


def _lookup_value(row: Dict[str, Any], *names: str) -> str:
    normalized = {str(key).strip().lower(): value for key, value in row.items()}
    for name in names:
        if name.lower() in normalized:
            value = normalized[name.lower()]
            return '' if value is None else str(value).strip()
    return ''


def import_institutions_from_excel(file_obj) -> int:
    rows = _read_rows(file_obj)
    created = 0
    with transaction.atomic():
        for row in rows:
            code = _lookup_value(row, 'code')
            name = _lookup_value(row, 'name')
            if not code or not name:
                continue
            Institution.objects.update_or_create(
                code=code,
                defaults={'name': name, 'is_active': True},
            )
            created += 1
    return created


def import_mentor_assignments_from_excel(file_obj) -> int:
    rows = _read_rows(file_obj)
    created = 0
    with transaction.atomic():
        for row in rows:
            institution_code = _lookup_value(row, 'institution_code', 'institution')
            mentor_username = _lookup_value(row, 'mentor_username', 'mentor')
            student_username = _lookup_value(row, 'student_username', 'student')
            if not institution_code or not mentor_username or not student_username:
                continue

            institution = Institution.objects.filter(code=institution_code).first()
            if institution is None:
                continue

            mentor = User.objects.filter(username=mentor_username).first()
            student = User.objects.filter(username=student_username).first()
            if mentor is None or student is None:
                continue

            mentor.institution = institution
            student.institution = institution
            mentor.save(update_fields=['institution'])
            student.save(update_fields=['institution'])

            role, _ = Role.objects.get_or_create(name=Role.MENTOR)
            UserRole.objects.get_or_create(user=mentor, role=role)

            assignment, was_created = MentorAssignment.objects.get_or_create(
                institution=institution,
                mentor=mentor,
                student=student,
            )
            if was_created:
                created += 1
    return created
