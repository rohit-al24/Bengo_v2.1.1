from io import BytesIO

from django.test import TestCase
from django.urls import reverse
from openpyxl import Workbook, load_workbook
from rest_framework.test import APIClient

from apps.accounts.models import Role, User, StudentProfile
from .excel_utils import build_template_bytes, import_institutions_from_excel, import_mentor_assignments_from_excel
from .models import Institution, MentorAssignment


class InstitutionModelTests(TestCase):
    def test_create_institution(self):
        institution = Institution.objects.create(code='ABC123', name='Sample Institute')
        self.assertEqual(str(institution), 'ABC123 - Sample Institute')


class InstitutionAssignmentApiTests(TestCase):
    def test_list_assignments_for_institution_admin(self):
        institution = Institution.objects.create(code='ABC123', name='Alpha Institute')
        admin = User.objects.create_user(username='institution_admin', email='admin@example.com', password='secret123')
        admin.institution = institution
        admin.save(update_fields=['institution'])
        role, _ = Role.objects.get_or_create(name=Role.INSTITUTIONAL_ADMIN)
        admin.roles.add(role)

        mentor = User.objects.create_user(username='mentor_user', email='mentor@example.com', password='secret123')
        student = User.objects.create_user(username='student_user', email='student@example.com', password='secret123')
        mentor.institution = institution
        student.institution = institution
        mentor.save(update_fields=['institution'])
        student.save(update_fields=['institution'])

        MentorAssignment.objects.create(institution=institution, mentor=mentor, student=student)

        client = APIClient()
        client.force_authenticate(user=admin)
        response = client.get(reverse('institution-assignments', kwargs={'institution_id': institution.id}))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.json()), 1)
        self.assertEqual(response.json()[0]['mentor_name'], mentor.username)


class InstitutionExcelImportTests(TestCase):
    def test_build_template_contains_expected_headers(self):
        template_bytes = build_template_bytes('institutions', ['code', 'name'])
        workbook = load_workbook(BytesIO(template_bytes))
        sheet = workbook.active

        self.assertEqual([sheet.cell(row=1, column=1).value, sheet.cell(row=1, column=2).value], ['code', 'name'])

    def test_import_institutions_from_excel_creates_records(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(['code', 'name'])
        sheet.append(['ABC123', 'Alpha Institute'])
        sheet.append(['XYZ789', 'Beta Institute'])

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        created = import_institutions_from_excel(buffer)

        self.assertEqual(created, 2)
        self.assertTrue(Institution.objects.filter(code='ABC123', name='Alpha Institute').exists())
        self.assertTrue(Institution.objects.filter(code='XYZ789', name='Beta Institute').exists())

    def test_import_mentor_assignments_from_excel_creates_assignments(self):
        institution = Institution.objects.create(code='ABC123', name='Alpha Institute')
        mentor = User.objects.create_user(username='mentor_user', email='mentor@example.com', password='secret123')
        student = User.objects.create_user(username='student_user', email='student@example.com', password='secret123')
        mentor.institution = institution
        student.institution = institution
        mentor.save(update_fields=['institution'])
        student.save(update_fields=['institution'])
        role, _ = Role.objects.get_or_create(name=Role.MENTOR)
        mentor.roles.add(role)

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(['institution_code', 'mentor_username', 'student_username'])
        sheet.append(['ABC123', 'mentor_user', 'student_user'])

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        created = import_mentor_assignments_from_excel(buffer)

        self.assertEqual(created, 1)
        self.assertTrue(MentorAssignment.objects.filter(institution=institution, mentor=mentor, student=student).exists())
