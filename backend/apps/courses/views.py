import random
import io
import re
import openpyxl
from django.http import HttpResponse
from django.utils import timezone
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status, permissions
from .models import Exam, Category, Lesson, StudyItem, QuestionBank, BankQuestion
from .serializers import (
    ExamSerializer, ExamDetailSerializer, ExamWriteSerializer,
    CategorySerializer, CategoryWriteSerializer,
    LessonSerializer, LessonWriteSerializer,
    StudyItemSerializer,
    QuestionBankSerializer, QuestionBankDetailSerializer,
    QuestionBankWriteSerializer, BankQuestionSerializer,
)
from apps.progress.models import UserExamUnlock, UserLessonProgress


# ─── Helpers ──────────────────────────────────────────────────────────────────

def is_admin(user):
    return user.is_authenticated and user.is_admin


def excel_template_response(filename, columns):
    """Return an Excel file with just the header row as download."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(columns)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


STUDY_COLUMNS = ['target', 'correct_answer', 'wrong_1', 'wrong_2', 'wrong_3']
EXAM_COLUMNS = ['target', 'correct_answer', 'wrong_1', 'wrong_2', 'wrong_3', 'wrong_4']
BANK_COLUMNS = STUDY_COLUMNS + ['wrong_4']


def _normalize_header(value):
    if value is None:
        return ''
    return re.sub(r'[^a-z0-9]+', '', str(value).strip().lower())


def _header_to_field(value):
    norm = _normalize_header(value)
    if 'target' in norm:
        return 'target'
    if 'correct' in norm and 'answer' in norm:
        return 'correct_answer'
    if 'wrong' in norm:
        match = re.search(r'wrong(?:answer)?(\d+)', norm)
        if match:
            return f'wrong_{match.group(1)}'
    if norm in {'exp1', 'exp2', 'exp3', 'exp4', 'exp5'}:
        return norm
    return None


def _build_header_map(header_row):
    mapping = {}
    for idx, cell in enumerate(header_row or []):
        field = _header_to_field(cell)
        if field:
            mapping[field] = idx
    return mapping


def _get_cell_value(row, index):
    if index is None or index < 0 or index >= len(row):
        return ''
    value = row[index]
    return '' if value is None else str(value).strip()


def _extract_row_values(row, lesson_type, is_topic, header_map=None):
    if is_topic:
        return {
            'target': _get_cell_value(row, header_map.get('target', 0) if header_map else 0),
            'exp1': _get_cell_value(row, header_map.get('exp1', 1) if header_map else 1),
            'exp2': _get_cell_value(row, header_map.get('exp2', 2) if header_map else 2),
            'exp3': _get_cell_value(row, header_map.get('exp3', 3) if header_map else 3),
            'exp4': _get_cell_value(row, header_map.get('exp4', 4) if header_map else 4),
            'exp5': _get_cell_value(row, header_map.get('exp5', 5) if header_map else 5),
        }

    values = {
        'target': _get_cell_value(row, header_map.get('target', 0) if header_map else 0),
        'correct_answer': _get_cell_value(row, header_map.get('correct_answer', 1) if header_map else 1),
        'wrong_1': _get_cell_value(row, header_map.get('wrong_1', 2) if header_map else 2),
        'wrong_2': _get_cell_value(row, header_map.get('wrong_2', 3) if header_map else 3),
        'wrong_3': _get_cell_value(row, header_map.get('wrong_3', 4) if header_map else 4),
    }
    if lesson_type == Lesson.EXAM:
        values['wrong_4'] = _get_cell_value(row, header_map.get('wrong_4', 5) if header_map else 5)
    return values


def lesson_is_visible_for_user(lesson, user):
    if not user or not getattr(user, 'is_authenticated', False):
        return lesson.rank_id is None
    if not lesson.rank_id:
        return True
    from apps.ranks.models import Rank, UserRankProgress

    active_rank = None
    progress = UserRankProgress.objects.filter(
        user=user, rank__exam=lesson.category.exam, is_current=True
    ).select_related('rank').first()
    if progress and progress.rank:
        active_rank = progress.rank
    else:
        active_rank = Rank.objects.filter(exam=lesson.category.exam).order_by('order').first()

    if active_rank and lesson.rank_id == active_rank.id:
        return True

    return UserRankProgress.objects.filter(
        user=user,
        rank_id=lesson.rank_id,
        is_current=True,
    ).exists()


# ─── Public / User views ──────────────────────────────────────────────────────

class ExamListView(APIView):
    """List all active exams, with unlock status for the current user."""
    def get(self, request):
        exams = Exam.objects.filter(is_active=True)
        data  = ExamSerializer(exams, many=True, context={'request': request}).data
        return Response(data)


class ExamDetailView(APIView):
    def get(self, request, pk):
        try:
            exam = Exam.objects.prefetch_related('categories__lessons').get(pk=pk, is_active=True)
        except Exam.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=404)
        return Response(ExamDetailSerializer(exam, context={'request': request}).data)


class UnlockExamView(APIView):
    def post(self, request, pk):
        try:
            exam = Exam.objects.get(pk=pk, is_active=True)
        except Exam.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=404)

        unlock, created = UserExamUnlock.objects.get_or_create(user=request.user, exam=exam)

        # Ensure the lowest rank for the unlocked exam is available to the user.
        first_rank = Rank.objects.filter(exam=exam).order_by('order').first()
        if first_rank:
            UserRankProgress.objects.get_or_create(
                user=request.user,
                rank=first_rank,
                defaults={'is_current': True, 'is_completed': False},
            )
        return Response({'unlocked': True, 'created': created})


class LessonStudyView(APIView):
    """Return study items for a lesson."""
    def get(self, request, pk):
        try:
            lesson = Lesson.objects.get(pk=pk, is_active=True)
        except Lesson.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=404)
        # Verify exam is unlocked
        if not UserExamUnlock.objects.filter(
            user=request.user, exam=lesson.category.exam
        ).exists():
            return Response({'detail': 'Exam not unlocked.'}, status=403)
        if not lesson_is_visible_for_user(lesson, request.user):
            return Response({'detail': 'Lesson not available for your rank.'}, status=403)
        items = lesson.study_items.all()
        return Response({
            'lesson':      LessonSerializer(lesson, context={'request': request}).data,
            'study_items': StudyItemSerializer(items, many=True).data,
        })


class LessonTestView(APIView):
    """Return shuffled quiz questions for a lesson's take-test."""
    def get(self, request, pk):
        try:
            lesson = Lesson.objects.get(pk=pk, is_active=True)
        except Lesson.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=404)
        if not UserExamUnlock.objects.filter(
            user=request.user, exam=lesson.category.exam
        ).exists():
            return Response({'detail': 'Exam not unlocked.'}, status=403)
        if not lesson_is_visible_for_user(lesson, request.user):
            return Response({'detail': 'Lesson not available for your rank.'}, status=403)

        questions = []

        if lesson.test_source == Lesson.FROM_STUDY:
            # Use study items as quiz questions
            items = list(lesson.study_items.all())
            random.shuffle(items)
            questions = StudyItemSerializer(items, many=True).data

        else:  # CUSTOM_BANK
            active_banks = lesson.question_banks.filter(is_active=True)
            all_q = list(BankQuestion.objects.filter(bank__in=active_banks))
            random.shuffle(all_q)
            count = active_banks.first().questions_count if active_banks.exists() else 40
            questions = BankQuestionSerializer(all_q[:count], many=True).data

        return Response({
            'lesson':    LessonSerializer(lesson, context={'request': request}).data,
            'questions': questions,
        })


class SubmitTestView(APIView):
    """Submit test answers and update lesson progress."""
    def post(self, request, pk):
        try:
            lesson = Lesson.objects.get(pk=pk, is_active=True)
        except Lesson.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=404)

        correct = int(request.data.get('correct', 0))
        total   = int(request.data.get('total', 1))
        score   = round((correct / total) * 100, 2) if total > 0 else 0.0
        
        # Get pass percentage from the lesson's assigned rank, otherwise current rank or first rank of the exam
        pass_pct = 70
        try:
            from apps.ranks.models import Rank, UserRankProgress
            if lesson.rank_id:
                pass_pct = lesson.rank.pass_percentage if lesson.rank else 70
            else:
                progress = UserRankProgress.objects.filter(
                    user=request.user,
                    rank__exam=lesson.category.exam,
                    is_current=True,
                ).select_related('rank').first()
                if progress and progress.rank:
                    pass_pct = progress.rank.pass_percentage
                else:
                    rank = Rank.objects.filter(exam=lesson.category.exam).order_by('order').first()
                    if rank:
                        pass_pct = rank.pass_percentage
        except Exception:
            pass

        passed  = score >= pass_pct

        prog, _ = UserLessonProgress.objects.get_or_create(
            user=request.user, lesson=lesson
        )
        prog.attempts     += 1
        prog.last_attempt  = timezone.now()
        if score > prog.best_score:
            prog.best_score = score
        if passed and not prog.is_completed:
            prog.is_completed = True
        prog.save()

        # Award XP
        if passed:
            xp_gain = 100
            request.user.xp += xp_gain
            request.user.save(update_fields=['xp'])
        else:
            xp_gain = 0

        if passed:
            try:
                from apps.ranks.models import Rank, UserRankProgress
                effective_rank = lesson.rank
                if effective_rank is None:
                    progress = UserRankProgress.objects.filter(
                        user=request.user,
                        rank__exam=lesson.category.exam,
                        is_current=True,
                    ).select_related('rank').first()
                    effective_rank = progress.rank if progress else None
                if effective_rank:
                    rank_progress, _ = UserRankProgress.objects.get_or_create(
                        user=request.user, rank=effective_rank
                    )
                    rank_progress.is_current = True
                    rank_progress.save(update_fields=['is_current'])
                    UserRankProgress.objects.filter(
                        user=request.user,
                        rank__exam=effective_rank.exam,
                    ).exclude(pk=rank_progress.pk).update(is_current=False)

                    if not rank_progress.is_completed:
                        lessons_for_rank = Lesson.objects.filter(rank=effective_rank, is_active=True)
                        completed_lessons = UserLessonProgress.objects.filter(
                            user=request.user,
                            lesson__in=lessons_for_rank,
                            is_completed=True,
                        ).count()
                        if lessons_for_rank.exists() and completed_lessons >= lessons_for_rank.count():
                            rank_progress.is_completed = True
                            rank_progress.completed_at = timezone.now()
                            rank_progress.save(update_fields=['is_completed', 'completed_at'])
            except Exception:
                pass

        return Response({
            'score':       score,
            'passed':      passed,
            'is_completed': prog.is_completed,
            'best_score':  prog.best_score,
            'xp_gained':   xp_gain,
            'pass_percentage': pass_pct,
        })


# ─── Admin views ──────────────────────────────────────────────────────────────

class AdminExamView(APIView):
    def get(self, request):
        if not is_admin(request.user):
            return Response({'detail': 'Forbidden.'}, status=403)
        exams = Exam.objects.prefetch_related('categories').all()
        return Response(ExamDetailSerializer(exams, many=True, context={'request': request}).data)

    def post(self, request):
        if not is_admin(request.user):
            return Response({'detail': 'Forbidden.'}, status=403)
        s = ExamWriteSerializer(data=request.data)
        if s.is_valid():
            return Response(ExamDetailSerializer(s.save(), context={'request': request}).data, status=201)
        return Response(s.errors, status=400)


class AdminExamDetailView(APIView):
    def _get(self, pk):
        try:
            return Exam.objects.get(pk=pk)
        except Exam.DoesNotExist:
            return None

    def get(self, request, pk):
        if not is_admin(request.user): return Response({'detail': 'Forbidden.'}, status=403)
        exam = self._get(pk)
        if not exam: return Response({'detail': 'Not found.'}, status=404)
        return Response(ExamDetailSerializer(exam, context={'request': request}).data)

    def patch(self, request, pk):
        if not is_admin(request.user): return Response({'detail': 'Forbidden.'}, status=403)
        exam = self._get(pk)
        if not exam: return Response({'detail': 'Not found.'}, status=404)
        s = ExamWriteSerializer(exam, data=request.data, partial=True)
        if s.is_valid():
            return Response(ExamDetailSerializer(s.save(), context={'request': request}).data)
        return Response(s.errors, status=400)

    def delete(self, request, pk):
        if not is_admin(request.user): return Response({'detail': 'Forbidden.'}, status=403)
        exam = self._get(pk)
        if not exam: return Response({'detail': 'Not found.'}, status=404)
        exam.delete()
        return Response(status=204)


class AdminCategoryView(APIView):
    def get(self, request):
        if not is_admin(request.user): return Response({'detail': 'Forbidden.'}, status=403)
        exam_id = request.query_params.get('exam')
        cats = Category.objects.filter(exam_id=exam_id) if exam_id else Category.objects.all()
        return Response(CategorySerializer(cats, many=True, context={'request': request}).data)

    def post(self, request):
        if not is_admin(request.user): return Response({'detail': 'Forbidden.'}, status=403)
        s = CategoryWriteSerializer(data=request.data)
        if s.is_valid():
            return Response(CategorySerializer(s.save(), context={'request': request}).data, status=201)
        return Response(s.errors, status=400)


class AdminCategoryDetailView(APIView):
    def _get(self, pk):
        try: return Category.objects.get(pk=pk)
        except Category.DoesNotExist: return None

    def get(self, request, pk):
        if not is_admin(request.user): return Response({'detail': 'Forbidden.'}, status=403)
        cat = self._get(pk)
        if not cat: return Response({'detail': 'Not found.'}, status=404)
        return Response(CategorySerializer(cat, context={'request': request}).data)

    def patch(self, request, pk):
        if not is_admin(request.user): return Response({'detail': 'Forbidden.'}, status=403)
        cat = self._get(pk)
        if not cat: return Response({'detail': 'Not found.'}, status=404)
        s = CategoryWriteSerializer(cat, data=request.data, partial=True)
        if s.is_valid():
            return Response(CategorySerializer(s.save(), context={'request': request}).data)
        return Response(s.errors, status=400)

    def delete(self, request, pk):
        if not is_admin(request.user): return Response({'detail': 'Forbidden.'}, status=403)
        cat = self._get(pk)
        if not cat: return Response({'detail': 'Not found.'}, status=404)
        cat.delete()
        return Response(status=204)



class AdminLessonView(APIView):
    def get(self, request):
        if not is_admin(request.user): return Response({'detail': 'Forbidden.'}, status=403)
        cat_id = request.query_params.get('category')
        lessons = Lesson.objects.filter(category_id=cat_id) if cat_id else Lesson.objects.all()
        return Response(LessonSerializer(lessons, many=True, context={'request': request}).data)

    def post(self, request):
        if not is_admin(request.user): return Response({'detail': 'Forbidden.'}, status=403)
        s = LessonWriteSerializer(data=request.data)
        if s.is_valid():
            return Response(LessonSerializer(s.save(), context={'request': request}).data, status=201)
        return Response(s.errors, status=400)


class AdminLessonDetailView(APIView):
    def _get(self, pk):
        try: return Lesson.objects.get(pk=pk)
        except Lesson.DoesNotExist: return None

    def patch(self, request, pk):
        if not is_admin(request.user): return Response({'detail': 'Forbidden.'}, status=403)
        lesson = self._get(pk)
        if not lesson: return Response({'detail': 'Not found.'}, status=404)
        s = LessonWriteSerializer(lesson, data=request.data, partial=True)
        if s.is_valid():
            return Response(LessonSerializer(s.save(), context={'request': request}).data)
        return Response(s.errors, status=400)

    def delete(self, request, pk):
        if not is_admin(request.user): return Response({'detail': 'Forbidden.'}, status=403)
        lesson = self._get(pk)
        if not lesson: return Response({'detail': 'Not found.'}, status=404)
        lesson.delete()
        return Response(status=204)


class AdminStudyImportView(APIView):
    """Import study items from Excel for a lesson."""

    FULL_ROW_COLS   = ['target', 'correct_answer', 'wrong_1', 'wrong_2', 'wrong_3']
    EXAM_COLS       = ['target', 'correct_answer', 'wrong_1', 'wrong_2', 'wrong_3', 'wrong_4']
    TOPIC_WISE_COLS = ['target', 'exp1', 'exp2', 'exp3', 'exp4', 'exp5']

    def _get_lesson(self, lesson_id):
        try:
            return Lesson.objects.select_related('category').get(pk=lesson_id)
        except Lesson.DoesNotExist:
            return None

    def get(self, request, lesson_id):
        """Download the Excel template — columns depend on category show_type."""
        if not is_admin(request.user): return Response({'detail': 'Forbidden.'}, status=403)
        lesson = self._get_lesson(lesson_id)
        if not lesson: return Response({'detail': 'Lesson not found.'}, status=404)
        if lesson.lesson_type == Lesson.EXAM:
            cols = self.EXAM_COLS
        elif lesson.category.show_type == 'topic_wise':
            cols = self.TOPIC_WISE_COLS
        else:
            cols = self.FULL_ROW_COLS
        return excel_template_response('study_template.xlsx', cols)

    def post(self, request, lesson_id):
        """Upload filled Excel to populate study items."""
        if not is_admin(request.user): return Response({'detail': 'Forbidden.'}, status=403)
        lesson = self._get_lesson(lesson_id)
        if not lesson: return Response({'detail': 'Lesson not found.'}, status=404)

        file = request.FILES.get('file')
        if not file:
            return Response({'detail': 'No file uploaded.'}, status=400)

        is_topic = lesson.category.show_type == 'topic_wise'

        wb  = openpyxl.load_workbook(file)
        ws  = wb.active
        rows = list(ws.iter_rows(values_only=True))

        header_map = {}
        data_rows = rows
        if rows and any(_header_to_field(cell) for cell in rows[0]):
            header_map = _build_header_map(rows[0])
            data_rows = rows[1:]

        created = 0
        for i, row in enumerate(data_rows):
            if not row or not any(_get_cell_value(row, idx) for idx in range(min(len(row), 6))):
                continue
            if is_topic:
                values = _extract_row_values(row, lesson.lesson_type, True, header_map)
                StudyItem.objects.create(
                    lesson = lesson,
                    target = values['target'],
                    exp1   = values['exp1'],
                    exp2   = values['exp2'],
                    exp3   = values['exp3'],
                    exp4   = values['exp4'],
                    exp5   = values['exp5'],
                    order  = i,
                )
            else:
                values = _extract_row_values(row, lesson.lesson_type, False, header_map)
                StudyItem.objects.create(
                    lesson         = lesson,
                    target         = values['target'],
                    correct_answer = values['correct_answer'],
                    wrong_1        = values['wrong_1'],
                    wrong_2        = values['wrong_2'],
                    wrong_3        = values['wrong_3'],
                    wrong_4        = values.get('wrong_4', ''),
                    order          = i,
                )
            created += 1

        return Response({'created': created, 'lesson': lesson.name, 'mode': 'topic_wise' if is_topic else 'full_row'})


class AdminQuestionBankView(APIView):
    def get(self, request, lesson_id):
        if not is_admin(request.user): return Response({'detail': 'Forbidden.'}, status=403)
        banks = QuestionBank.objects.filter(lesson_id=lesson_id)
        return Response(QuestionBankSerializer(banks, many=True).data)

    def post(self, request, lesson_id):
        if not is_admin(request.user): return Response({'detail': 'Forbidden.'}, status=403)
        data = {**request.data, 'lesson': lesson_id}
        s = QuestionBankWriteSerializer(data=data)
        if s.is_valid():
            return Response(QuestionBankSerializer(s.save()).data, status=201)
        return Response(s.errors, status=400)


class AdminQuestionBankDetailView(APIView):
    def _get(self, pk):
        try: return QuestionBank.objects.get(pk=pk)
        except QuestionBank.DoesNotExist: return None

    def get(self, request, pk):
        if not is_admin(request.user): return Response({'detail': 'Forbidden.'}, status=403)
        bank = self._get(pk)
        if not bank: return Response({'detail': 'Not found.'}, status=404)
        return Response(QuestionBankDetailSerializer(bank).data)

    def patch(self, request, pk):
        if not is_admin(request.user): return Response({'detail': 'Forbidden.'}, status=403)
        bank = self._get(pk)
        if not bank: return Response({'detail': 'Not found.'}, status=404)
        s = QuestionBankWriteSerializer(bank, data=request.data, partial=True)
        if s.is_valid():
            return Response(QuestionBankSerializer(s.save()).data)
        return Response(s.errors, status=400)

    def delete(self, request, pk):
        if not is_admin(request.user): return Response({'detail': 'Forbidden.'}, status=403)
        bank = self._get(pk)
        if not bank: return Response({'detail': 'Not found.'}, status=404)
        bank.delete()
        return Response(status=204)


class AdminBankImportView(APIView):
    """Import questions into a question bank from Excel."""

    def get(self, request, bank_id):
        if not is_admin(request.user): return Response({'detail': 'Forbidden.'}, status=403)
        return excel_template_response('bank_template.xlsx', BANK_COLUMNS)

    def post(self, request, bank_id):
        if not is_admin(request.user): return Response({'detail': 'Forbidden.'}, status=403)
        try:
            bank = QuestionBank.objects.get(pk=bank_id)
        except QuestionBank.DoesNotExist:
            return Response({'detail': 'Bank not found.'}, status=404)

        file = request.FILES.get('file')
        if not file:
            return Response({'detail': 'No file uploaded.'}, status=400)

        wb  = openpyxl.load_workbook(file)
        ws  = wb.active
        created = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]: continue
            BankQuestion.objects.create(
                bank           = bank,
                target         = str(row[0] or '').strip(),
                correct_answer = str(row[1] or '').strip(),
                wrong_1        = str(row[2] or '').strip(),
                wrong_2        = str(row[3] or '').strip(),
                wrong_3        = str(row[4] or '').strip(),
                wrong_4        = str(row[5] or '').strip() if len(row) > 5 else '',
            )
            created += 1
        return Response({'created': created, 'bank': bank.title})


class AdminStudyItemsBulkDeleteView(APIView):
    """Bulk delete study items."""

    def post(self, request):
        if not is_admin(request.user):
            return Response({'detail': 'Forbidden.'}, status=403)

        ids = request.data.get('ids', [])
        if not ids:
            return Response({'detail': 'No IDs provided.'}, status=400)

        deleted_count, _ = StudyItem.objects.filter(id__in=ids).delete()
        return Response({'deleted': deleted_count})
